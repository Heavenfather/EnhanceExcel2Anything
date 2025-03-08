import openpyxl
import re
import json
import core.utils.utils
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from core.utils.type_system import TypeSystem
from typing import Dict, Any, List
from core.models import SheetConfig, FieldMeta
from core.utils.exceptions import TypeCastError, JsonTypeCastError, ConfigError
from decimal import Decimal, InvalidOperation
from json.decoder import JSONDecodeError
from dataclasses import replace


class ExcelProcessor:
    def __init__(self, type_system: TypeSystem):
        self.type_system = type_system

    def process_workbook(self, file_path: str) -> List[SheetConfig]:
        """
        处理Excel文件
        :param file_path: 指定Excel文件路径
        """
        configs = []
        wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=False, keep_links=False)
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('#'):
                continue
            sheet = wb[sheet_name]
            config = self.__process_sheet(sheet, file_path)
            configs.append(config)
        wb.close()
        return configs

    def __process_sheet(self, sheet, file_path) -> SheetConfig:
        """处理每个工作簿"""
        if sheet.max_row < 4 or sheet.max_column < 2:
            raise RuntimeError(f'表格格式错误,请检查表格是否正确 [{sheet.title}]')

        # 解析导出名称 A1位置
        export_name = sheet['A1'].value
        if not export_name:
            raise ValueError(
                f'导出名称为空 [{file_path}:{sheet.title}],A1位置必须填写导出名称,如无需导出Sheet名称则填写#开头')

        fields = {}
        for col_idx in range(2, sheet.max_column + 1):  # 跳过第一列
            header = sheet.cell(row=1, column=col_idx).value
            if header is None:
                continue
            if not core.utils.utils.validate_str_legal(header):
                raise ValueError(f'字段名称非法 [{file_path}:{sheet.title}]-->{header}')
            # 解析字段元数据
            field_type = sheet.cell(row=2, column=col_idx).value
            if field_type is None:
                continue
            # 校验配置类型
            if not self.type_system.is_support_type(field_type):
                raise ValueError(f'不受支持的字段类型 [{file_path}:{sheet.title}]-->{field_type}')
            checker = sheet.cell(row=3, column=col_idx).value or ''
            comment = sheet.cell(row=4, column=col_idx).value or ''

            fields[header] = FieldMeta(
                name=header.strip(),
                type=field_type,
                checks=[t.strip() for t in checker.split(';')],
                comment=comment,
                col_index=col_idx,
                is_ignored=header.startswith('#') or field_type is None
            )

        # 多线程解析数据行
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = []
            for row_idx in range(5, sheet.max_row + 1):
                futures.append(executor.submit(self.__process_row, sheet, row_idx, fields))

            data_rows = [f.result() for f in futures if f.result() is not None]

        return SheetConfig(
            export_name=export_name.strip(),
            fields=fields,
            rows_values=data_rows,
            sheets=[sheet.title],
            source_file=file_path,
            source_file_md5=0
        )

    def __process_row(self, sheet, row_idx, fields):
        row_values = {}
        for field in fields.values():
            if field.is_ignored:
                continue
            cell = sheet.cell(row=row_idx, column=field.col_index)
            try:
                parsed_value = self.__parse_cell_value(cell.value, field)
                row_values[field.name] = parsed_value
            except Exception as e:
                raise ConfigError(f'解析行数据失败 [错误Sheet:{sheet.title} 行:{row_idx} 列:{field.col_index}]: {e}')
        return row_values

    def __parse_cell_value(self, raw_value, field):
        # 处理默认值
        if raw_value is None:
            return self.__get_default_value(field)

        # 类型转换逻辑
        return self.__cast_value(raw_value, field)

    def __get_default_value(self, field: FieldMeta):
        """
        获取字段的默认值 处理优先级
        1.配置字段标签中的Default值
        2.自定义类型的default值
        3.数据类型默认值
        """

        # 解析字段标签中的默认值
        for tag in field.checks:
            if tag.startswith('Default:'):
                default_str = tag.split(':', 1)[1].strip()
                return self.__cast_value(default_str, field)

        try:
            return self.type_system.get_default_value(field.type)
        except KeyError:
            return self.__infer_default_by_type(field.type)

    def __infer_default_by_type(self, data_type: str) -> Any:
        """类型推断后备方案"""
        if data_type.lower().startswith("list<"):
            return []
        if data_type.lower().startswith("map<"):
            return {}
        if data_type.startswith("datetime"):
            return 0
        return None

    def __cast_value(self, raw_value: Any, field: FieldMeta) -> Any:
        """
        类型转换核心逻辑，支持以下类型：
        - 基础类型：int/long/float/bool/string等
        - 复合类型：list/map
        - 日期类型:datetime(YYYY-MM-DD or YYYY-MM-DD HH:MM)
        - 自定义类型：struct/enum/class
        """
        if raw_value is None:
            return None

        # 取出首尾空格
        if isinstance(raw_value, str):
            raw_value = raw_value.strip()
            if not raw_value:
                return None

        # 递归处理泛型
        data_type = field.type
        if data_type.lower().startswith("list<"):
            return self.__cast_list(raw_value, field)
        if data_type.lower().startswith("map<"):
            return self.__cast_dict(raw_value, field)

        # 自定义类型处理
        if self.type_system.is_custom_support_type(data_type):
            return self.__cast_custom_type(raw_value, field)

        return self.__cast_basic_type(raw_value, field)

    def __cast_list(self, raw_value: Any, field: FieldMeta) -> List:
        """list类型转换"""
        if not raw_value:
            return []

        inner_type = field.type[5:-1]
        # 默认使用 | 分割
        seps = r"[|]"
        # 解析自定义分隔符 标签ListSeparator
        if field.checks:
            for tag in field.checks:
                if tag.startswith('ListSeparator:'):
                    seps = tag.split(':', 1)[1]
                    break
        items = re.split(seps, str(raw_value))
        return [self.__cast_value(item.strip(), replace(field, type=inner_type))
                for item in items if item.strip()
                ]

    def __cast_dict(self, raw_value: str, field: FieldMeta):
        """map类型转换
        配置表配置格式：
            key:value
        """
        # 解析键值类型 map<key,value>
        key_type, value_type = field.type[4:-1].split(',', 1)
        result = {}
        seps = r"[|]"
        # 解析自定义分隔符 标签MapSeparator 支持map内可以嵌套多个键值，省掉又包一层list
        if field.checks:
            for tag in field.checks:
                if tag.startswith('MapSeparator:'):
                    seps = f'[{tag.split(":", 1)[1]}]'
                    break
        items = re.split(seps, str(raw_value))
        for pair in items:
            if not pair:
                continue
            key, value = pair.split(':', 1)
            cast_key = self.__cast_value(key.strip(), replace(field, type=key_type))
            cast_value = self.__cast_value(value.strip(), replace(field, type=value_type))
            result[cast_key] = cast_value
        return result

    def __cast_custom_type(self, raw_value: str, field: FieldMeta) -> Any:
        """处理自定义类型"""
        type_defs = self.type_system.get_type_definition(field.type)
        if type_defs['type'] == 'enum':
            if raw_value not in type_defs['fields']:
                raise TypeCastError(
                    f"枚举值 '{raw_value}' 不在允许范围内 {type_defs['values']}")
            return str(raw_value)

        if type_defs["type"] in ("struct", "class"):
            return self.__parse_struct_json(raw_value, type_defs, field)

        # 其它类型处理...TODO

    def __parse_struct_json(self, raw_value: str, type_defs: dict, field: FieldMeta):
        """解析并验证JSON格式的结构体数据"""
        # 解析并转换Json格式，允许用户不用加上双引号
        formatted_json = re.sub(
            r'([{,]\s*)([a-zA-Z_]\w*)(\s*:)',
            r'\1"\2"\3',
            raw_value
        )
        # 处理布尔值和null 由于Python的json库不支持，需要特殊处理
        formatted_json = re.sub(
            r':\s*(true|false|null)\b',
            lambda m: f': {m.group(1).lower()}',
            formatted_json,
            flags=re.IGNORECASE
        )

        try:
            json_data = json.loads(formatted_json)
        except JSONDecodeError as e:
            raise JsonTypeCastError(
                f"无效的JSON格式: {e.msg}",
                original_value=raw_value,
                error_position=f"第{e.lineno}行，列{e.colno}"
            ) from e

        # 验证字段完整性
        required_fields = set(type_defs['fields'].keys())
        provided_fields = set(json_data.keys())
        if missing := required_fields - provided_fields:
            raise JsonTypeCastError(
                f"缺少必要字段: {', '.join(missing)}",
                original_value=raw_value,
                missing_fields=list(missing)
            )

        # 检查多余字段
        if extra := provided_fields - required_fields:
            raise JsonTypeCastError(
                f"存在多余字段: {', '.join(extra)}",
                original_value=raw_value,
                extra_fields=list(extra)
            )

        # 递归转换每个字段
        result = {}
        for field_name, field_type in type_defs['fields'].items():
            raw_field_value = json_data.get(field_name)

            try:
                result[field_name] = self.__cast_value(raw_field_value, replace(field, type=field_type, checks=[]))
            except JsonTypeCastError as e:
                e.add_context(f"字段 '{field_name}'")
                raise
            except Exception as e:
                raise JsonTypeCastError(
                    f"字段 '{field_name}' 转换失败: {e}",
                    original_value=raw_value,
                    data_type=field_type
                )
        return result

    def __cast_basic_type(self, raw_value: Any, field: FieldMeta) -> Any:
        """基础类型转换"""
        data_type = field.type
        try:
            if data_type in ['int', 'long', 'byte', 'sbyte', 'short', 'uint', 'ulong', 'ushort']:
                return self.__cast_integer(raw_value, data_type)
            if data_type in ['float', 'double', 'decimal']:
                return self.__cast_float(raw_value, data_type)
            if data_type == 'bool':
                return self.__cast_boolean(raw_value)
            if data_type == 'datetime':
                return self.__cast_datetime(raw_value, field)
            if data_type == 'string':
                return str(raw_value) if raw_value is not None else ''
        except (ValueError, TypeError) as e:
            error_detail = f"值 '{raw_value}' ({type(raw_value).__name__}) -> {data_type}"
            raise TypeCastError(f"基础类型转换失败: {error_detail}") from e

    def __cast_datetime(self, raw_value: str, field: FieldMeta):
        """解析日期配置格式
            - 返回时间戳
        """
        if not raw_value:
            return 0
        if isinstance(raw_value, datetime):
            return int(raw_value.timestamp())
        if not isinstance(raw_value, str):
            return 0
        # 获取日期格式，默认 %Y/%m/%d %H:%M:%S
        date_format = '%Y/%m/%d %H:%M:%S'
        # 解析自定义日期格式 标签DateFormat:
        for check in field.checks:
            if check.startswith('DateFormat:'):
                date_format = check.split(':', 1)[1]
                break
        # 配置格式校验
        if not re.match(self.__build_datetime_regex_pattern(date_format), str(raw_value)):
            raise TypeCastError(f"日期格式错误: {raw_value},要求格式: {date_format}")

        try:
            dt = datetime.strptime(str(raw_value), date_format)
            return int(dt.timestamp())
        except Exception as e:
            raise TypeCastError(f"无效日期: {raw_value}") from e

    def __build_datetime_regex_pattern(self, date_format: str) -> str:
        """转换正则表达式"""
        format_map = {
            "%Y": r"\d{4}",  # 年
            "%m": r"\d{1,2}",  # 月（允许1-2位）
            "%d": r"\d{1,2}",  # 日（允许1-2位）
            "%H": r"\d{2}",  # 小时（严格两位）
            "%M": r"\d{2}",  # 分（严格两位）
            "%S": r"\d{2}",  # 秒（严格两位）
        }

        # 拆分格式字符串为动态部分和静态分隔符
        pattern = []
        i = 0
        while i < len(date_format):
            if date_format[i] == '%' and i + 1 < len(date_format):
                specifier = date_format[i:i + 2]
                pattern.append(format_map.get(specifier, specifier))
                i += 2
            else:
                # 转义静态字符并保留原分隔符
                pattern.append(re.escape(date_format[i]))
                i += 1

        return f"^{''.join(pattern)}$"

    def __cast_integer(self, value: Any, type_name: str) -> int:
        """处理所有整数类型转换(支持科学计数)"""
        original_value = value

        # 预处理科学计数法
        if isinstance(value, str):
            value = value.strip().lower()
            if 'e' in value:
                try:
                    # 先转换为浮点数处理科学计数法
                    float_value = float(value)
                    if not float_value.is_integer():
                        raise ValueError(f'科学计数法数值 {original_value} 不是整数')
                    value = float_value
                except ValueError:
                    raise ValueError(f'科学计数法数值 {original_value} 无效')

        # 统一转换为整数
        try:
            if isinstance(value, float):
                if not value.is_integer():
                    raise ValueError(f'浮点数 {original_value} 不是整数')
                int_value = int(value)
            else:
                int_value = int(value)
        except ValueError:
            raise ValueError(f'值 {original_value} 无效')

        # 范围校验
        bits_map = {
            'byte': 8,
            'sbyte': 8,
            'short': 16,
            'ushort': 16,
            'int': 32,
            'uint': (32, False),
            'long': 64,
            'ulong': (64, False)
        }
        if type_name in bits_map:
            bits = bits_map[type_name]
            if isinstance(bits, tuple):
                bits, signed = bits
            else:
                signed = not type_name.startswith('u')

            min_val = - (2 ** (bits - 1)) if signed else 0
            max_val = (2 ** (bits - (1 if signed else 0))) - 1

            if not (min_val <= int_value <= max_val):
                raise OverflowError(f"值 {int_value} 超出 {type_name} 范围")

        return int_value

    def __cast_float(self, value: Any, type_name: str) -> float:
        """处理浮点类型转换（支持科学计数法、逗号分隔符等）"""
        try:
            # 统一转为字符串处理
            original = str(value).strip() if isinstance(value, str) else str(value)

            # 步骤 1：基础清理（保留数字、科学计数符号、分隔符、正负号）
            cleaned = re.sub(r"[^\d.,eE+-]", "", original)
            if not cleaned:
                return 0.0

            # 步骤 2：符号合法性检查
            sign_chars = sum(1 for c in cleaned if c in '+-')
            if sign_chars > 2 or (sign_chars > 1 and 'e' not in cleaned.lower()):
                raise TypeCastError(f"符号错误 [{type_name}]: {original}")

            # 步骤 3：类型分支处理
            if type_name == 'decimal':
                return self.__cast_decimal(cleaned)

            # 步骤 4：预处理后转换为浮点数
            normalized = self.__normalize_separators(cleaned)
            return float(normalized)

        except (ValueError, TypeError, InvalidOperation) as e:
            error_msg = f"数值转换失败 [{type_name}]: {original} -> {cleaned}"
            raise TypeCastError(error_msg) from e

    def __cast_decimal(self, value_str: str) -> Decimal:
        """高精度十进制转换（支持配置精度）"""
        try:
            # 移除多余的小数点（如 "12.34.56" -> 报错）
            if value_str.count('.') > 1:
                raise ValueError(f"多个小数点 : {value_str}")

            # 自动补全不完整小数（如 "123." -> 123.0）
            if value_str.endswith('.'):
                value_str += '0'

            return Decimal(value_str).normalize()
        except InvalidOperation as e:
            # 处理特殊值（如 NaN, Infinity）
            if value_str.lower() in {'nan', 'inf', 'infinity'}:
                return Decimal('NaN')
            raise TypeCastError(f"Decimal转换失败: {value_str}") from e

    def __normalize_separators(self, s: str) -> str:
        """智能处理数字分隔符（支持欧美格式）"""
        # 分离科学计数法部分
        if 'e' in s or 'E' in s:
            base_part, exp_part = re.split(r'[eE]', s, 1)
            return f"{self.__process_base_part(base_part)}E{exp_part}"
        return self.__process_base_part(s)

    def __process_base_part(self, s: str) -> str:
        """处理基数部分的分隔符"""
        # 统计所有分隔符
        separators = [i for i, c in enumerate(s) if c in ',.']

        # 没有分隔符直接返回
        if not separators:
            return s

        # 以最后一个分隔符作为小数点
        last_sep_pos = separators[-1]
        parts = []
        for i, c in enumerate(s):
            if c in ',.':
                if i == last_sep_pos:
                    parts.append('.')
                else:  # 移除非小数点分隔符
                    continue
            else:
                parts.append(c)
        return ''.join(parts)

    def __cast_boolean(self, value: Any) -> bool:
        """处理布尔类型转换"""
        if isinstance(value, bool):
            return value
        str_value = str(value).lower()
        if str_value in {'true', '1', 'yes', 'y'}:
            return True
        if str_value in {'false', '0', 'no', 'n'}:
            return False
        raise ValueError(f"无法解析的布尔值: {value}")
